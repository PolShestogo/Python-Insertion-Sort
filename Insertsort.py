import random
import timeit
from openpyxl import load_workbook
from openpyxl import Workbook

def insertSort(arr):
        for i in range (1,len(arr)):
                t=int(arr[i])
                j=i-1
                while(j>=0 and t<int(arr[j])):
                        arr[j+1]=arr[j]
                        j=j-1
                arr[j+1]=t

def exeltable(values):
        workbook = load_workbook(filename="123.xlsx")
        #workbook=Workbook()        
        sheet = workbook.active
        for i in range (1,len(values)):
                sheet["A"+str(i)] = values[i][0]
                sheet["B"+str(i)] = values[i][1]
        workbook.save(filename="123.xlsx")

def sort(n, bool=False):
        max=1000
        min=-1000
        nmax=100000              
        file=open('input.txt','w+')
        file.write(str(random.randint(min,max)))
        for i in range (1, int(n)-1):
                file.write(' ')
                file.write(str(random.randint(min,max)))       
        file.close()
        file=open('input.txt','r')
        arr=list((file.read()).split())
        if bool :
                print('Исходный массив:', str(arr))
        start = timeit.default_timer()
        insertSort(arr)
        end= timeit.default_timer()
        if bool :
                print('Отсортированный массив: ', arr)
        sorted=open('sorted.txt','w+')
        sorted.write(str(arr))
        t=end-start
        if bool :
                print("Время сортировки: ",t )
        print(n)
        return(n,t)


def main():  
       #n= input("Введите количество элементов сортируемого массива: ")
       values=[[]]
       j=100
       n=50
       for i in range(1,j):
                values.append(sort(int(n)*i))
       exeltable(values)
       


        
if __name__ == '__main__':
    main()